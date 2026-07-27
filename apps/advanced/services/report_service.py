"""
سرویس تولید گزارشات (Excel/CSV)
"""
import io
import csv
from datetime import datetime
from django.core.files.base import ContentFile
from django.utils import timezone
from django.db.models import Sum, Count, Avg

from apps.advanced.models import Report


class ReportService:
    """سرویس تولید گزارشات"""

    @classmethod
    def create_report(cls, user, report_type, format_type, filters):
        """
        ایجاد درخواست گزارش
        """
        report = Report.objects.create(
            user=user,
            report_type=report_type,
            format=format_type,
            filters=filters,
        )

        # تولید همزمان (برای ساده‌سازی)
        cls._generate_report(report)

        return report

    @classmethod
    def _generate_report(cls, report):
        """تولید فایل گزارش"""
        try:
            if report.report_type == Report.Type.TRANSACTIONS:
                data = cls._get_transactions_data(report.user, report.filters)
            elif report.report_type == Report.Type.APPOINTMENTS:
                data = cls._get_appointments_data(report.user, report.filters)
            elif report.report_type == Report.Type.REVIEWS:
                data = cls._get_reviews_data(report.user, report.filters)
            else:
                raise ValueError(f'نوع گزارش پشتیبانی نمی‌شود: {report.report_type}')

            # تولید فایل
            if report.format == Report.Format.CSV:
                file_content = cls._generate_csv(data)
                filename = f'report_{report.report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            else:
                file_content = cls._generate_excel(data, report.report_type)
                filename = f'report_{report.report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

            # ذخیره فایل
            report.file.save(filename, ContentFile(file_content))
            report.records_count = len(data)
            report.file_size = len(file_content)
            report.is_ready = True
            report.completed_at = timezone.now()
            report.expires_at = timezone.now() + timezone.timedelta(days=7)
            report.save()

        except Exception as e:
            report.error_message = str(e)
            report.save()

    @classmethod
    def _get_transactions_data(cls, user, filters):
        """دریافت داده‌های تراکنش‌ها"""
        from apps.payments.models import Transaction

        qs = Transaction.objects.filter(business__owner=user)

        # فیلترها
        if filters.get('date_from'):
            qs = qs.filter(created_at__date__gte=filters['date_from'])
        if filters.get('date_to'):
            qs = qs.filter(created_at__date__lte=filters['date_to'])
        if filters.get('status'):
            qs = qs.filter(status=filters['status'])
        if filters.get('type'):
            qs = qs.filter(type=filters['type'])

        qs = qs.select_related('user', 'business', 'appointment').order_by('-created_at')

        return [
            {
                'کد پیگیری': tx.tracking_code,
                'تاریخ': tx.created_at.strftime('%Y/%m/%d %H:%M'),
                'نام مشتری': tx.user.full_name or tx.user.phone,
                'مبلغ (تومان)': tx.amount,
                'کارمزد (تومان)': tx.commission_amount,
                'مبلغ خالص (تومان)': tx.net_amount,
                'وضعیت': tx.get_status_display(),
                'نوع': tx.get_type_display(),
                'درگاه': tx.gateway or '—',
            }
            for tx in qs
        ]

    @classmethod
    def _get_appointments_data(cls, user, filters):
        """دریافت داده‌های نوبت‌ها"""
        from apps.bookings.models import Appointment

        qs = Appointment.objects.filter(business__owner=user)

        if filters.get('date_from'):
            qs = qs.filter(date__gte=filters['date_from'])
        if filters.get('date_to'):
            qs = qs.filter(date__lte=filters['date_to'])
        if filters.get('status'):
            qs = qs.filter(status=filters['status'])

        qs = qs.select_related(
            'customer', 'service', 'employee'
        ).order_by('-date', '-time')

        return [
            {
                'تاریخ': apt.date.strftime('%Y/%m/%d'),
                'ساعت': apt.time.strftime('%H:%M'),
                'نام مشتری': apt.customer.full_name or apt.customer.phone,
                'شماره مشتری': apt.customer.phone,
                'خدمت': apt.service.name,
                'کارمند': apt.employee.name if apt.employee else '—',
                'مبلغ کل (تومان)': apt.final_price,
                'بیعانه (تومان)': apt.deposit_amount,
                'وضعیت': apt.get_status_display(),
            }
            for apt in qs
        ]

    @classmethod
    def _get_reviews_data(cls, user, filters):
        """دریافت داده‌های نظرات"""
        from apps.reviews.models import Review

        qs = Review.objects.filter(business__owner=user)

        if filters.get('date_from'):
            qs = qs.filter(created_at__date__gte=filters['date_from'])
        if filters.get('date_to'):
            qs = qs.filter(created_at__date__lte=filters['date_to'])
        if filters.get('rating'):
            qs = qs.filter(rating=filters['rating'])

        qs = qs.select_related(
            'customer', 'service', 'business'
        ).order_by('-created_at')

        return [
            {
                'تاریخ': review.created_at.strftime('%Y/%m/%d'),
                'نام مشتری': review.customer.full_name or review.customer.phone,
                'خدمت': review.service.name if review.service else '—',
                'امتیاز': review.rating,
                'نظر': review.comment,
                'وضعیت': 'تایید شده' if review.is_approved else 'در انتظار',
            }
            for review in qs
        ]

    @classmethod
    def _generate_csv(cls, data):
        """تولید CSV"""
        if not data:
            return b''

        output = io.StringIO()
        # نوشتن با BOM برای پشتیبانی از فارسی در Excel
        output.write('\ufeff')

        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

        return output.getvalue().encode('utf-8-sig')

    @classmethod
    def _generate_excel(cls, data, report_type):
        """تولید Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            # اگر openpyxl نصب نیست، به CSV برگرد
            return cls._generate_csv(data)

        if not data:
            return b''

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Report'

        # Header Style
        header_font = Font(name='Tahoma', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(
            start_color='A88B7D',
            end_color='A88B7D',
            fill_type='solid',
        )
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data
        for row, record in enumerate(data, 2):
            for col, key in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=record[key])
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-adjust columns
        for col in range(1, len(headers) + 1):
            max_length = max(
                len(str(ws.cell(row=row, column=col).value or ''))
                for row in range(1, len(data) + 2)
            )
            ws.column_dimensions[chr(64 + col)].width = min(max_length + 2, 30)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()