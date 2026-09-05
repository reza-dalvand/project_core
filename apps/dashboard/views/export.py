# apps/dashboard/views/export.py
"""
✅ فاز ۳: خروجی اکسل/CSV — رفع ۳ باگ
- ۳.۸.۱: اطلاع‌رسانی در صورت فال‌بک به CSV
- ۳.۸.۲: محدودیت حجم خروجی
- ۳.۸.۳: هندل خطاها در _get_nested_field
"""
import csv
import io
import logging
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import redirect
from django.utils import timezone
from apps.accounts.models import User
from apps.appointments.models import Appointment
from apps.businesses.models import Business
from apps.dashboard.decorators import admin_login_required
from apps.payments.models import Settlement, Transaction
from apps.support.models import SupportTicket

logger = logging.getLogger(__name__)

# ✅ FIX ۳.۸.۲: حداکثر تعداد رکورد در خروجی
MAX_EXPORT_ROWS = 50000

# ─── مدل‌های پشتیبانی شده برای خروجی ───
EXPORT_MODELS = {
    'users': {
        'model': User,
        'filename': 'users_export',
        'headers': [
            'شناسه', 'شماره تلفن', 'نام', 'نام خانوادگی',
            'ایمیل', 'وضعیت فعال', 'وضعیت تایید', 'تاریخ عضویت',
        ],
        'fields': [
            'id', 'phone', 'first_name', 'last_name',
            'email', 'is_active', 'is_verified', 'date_joined',
        ],
        'roles': ['super_admin', 'app_admin'],
    },
    'businesses': {
        'model': Business,
        'filename': 'businesses_export',
        'headers': [
            'شناسه', 'نام', 'مالک', 'شهر', 'استان',
            'وضعیت', 'امتیاز', 'تعداد نظرات', 'تاریخ ایجاد',
        ],
        'fields': [
            'id', 'name', 'owner__phone', 'city__name',
            'province__name', 'status', 'rating',
            'reviews_count', 'created_at',
        ],
        'roles': ['super_admin', 'app_admin'],
    },
    'transactions': {
        'model': Transaction,
        'filename': 'transactions_export',
        'headers': [
            'شناسه', 'کد پیگیری', 'مشتری', 'کسب‌وکار',
            'نوع', 'مبلغ', 'کارمزد', 'وضعیت', 'تاریخ ایجاد',
        ],
        'fields': [
            'id', 'tracking_code', 'customer__phone',
            'business__name', 'type', 'amount',
            'app_fee', 'status', 'created_at',
        ],
        'roles': ['super_admin', 'financial_admin'],
    },
    'settlements': {
        'model': Settlement,
        'filename': 'settlements_export',
        'headers': [
            'شناسه', 'کسب‌وکار', 'مبلغ', 'وضعیت',
            'شبا', 'بانک', 'تاریخ ایجاد',
        ],
        'fields': [
            'id', 'business__name', 'amount', 'status',
            'bank_sheba', 'bank_name', 'created_at',
        ],
        'roles': ['super_admin', 'financial_admin'],
    },
    'tickets': {
        'model': SupportTicket,
        'filename': 'tickets_export',
        'headers': [
            'شناسه', 'کاربر', 'موضوع', 'وضعیت',
            'اولویت', 'تاریخ ایجاد',
        ],
        'fields': [
            'id', 'user__phone', 'subject', 'status',
            'priority', 'created_at',
        ],
        'roles': ['super_admin', 'support_admin'],
    },
    'appointments': {
        'model': Appointment,
        'filename': 'appointments_export',
        'headers': [
            'شناسه', 'مشتری', 'کسب‌وکار', 'خدمت',
            'تاریخ', 'ساعت', 'وضعیت', 'مبلغ کل',
        ],
        'fields': [
            'id', 'customer__phone', 'business__name',
            'service__name', 'date_key', 'time_slot',
            'status', 'total_price',
        ],
        'roles': ['super_admin', 'app_admin'],
    },
}


@admin_login_required
def export_view(request):
    """
    خروجی اکسل/CSV
    GET /dashboard/export/?model=users&format=csv
    GET /dashboard/export/?model=users&format=excel
    """
    model_name = request.GET.get('model', '')
    export_format = request.GET.get('format', 'csv')

    # ─── اعتبارسنجی مدل ───
    if model_name not in EXPORT_MODELS:
        messages.error(request, 'نوع داده برای خروجی نامعتبر است.')
        return redirect('dashboard:home')

    model_config = EXPORT_MODELS[model_name]
    model = model_config['model']

    # ─── بررسی دسترسی نقش ───
    user_role = request.session.get('dashboard_role', '')
    if user_role not in model_config['roles']:
        messages.error(request, 'شما به خروجی این داده‌ها دسترسی ندارید.')
        return redirect('dashboard:home')

    # ─── فیلترهای اختیاری ───
    queryset = model.objects.all()

    search = request.GET.get('search', '')
    if search:
        text_fields = [
            field
            for field in model_config['fields']
            if not field.startswith('created_at')
            and not field.startswith('date_joined')
        ]
        q_filters = Q()
        for field in text_fields:
            if '__' in field:
                q_filters |= Q(**{f'{field}__icontains': search})
        if q_filters:
            queryset = queryset.filter(q_filters)

    # ✅ FIX ۳.۸.۲: بررسی تعداد کل رکوردها
    total_count = queryset.count()
    if total_count > MAX_EXPORT_ROWS:
        messages.warning(
            request,
            f'تعداد رکوردها ({total_count:,}) بیشتر از '
            f'حداکثر مجاز ({MAX_EXPORT_ROWS:,}) است. '
            f'فقط {MAX_EXPORT_ROWS:,} رکورد اول خروجی گرفته می‌شود.'
        )
        queryset = queryset[:MAX_EXPORT_ROWS]

    # ─── تولید خروجی ───
    try:
        if export_format == 'excel':
            return _generate_excel_export(request, queryset, model_config)
        return _generate_csv_export(request, queryset, model_config)
    except Exception as e:
        logger.error(f'Export error: {e}', exc_info=True)
        messages.error(request, 'خطا در تولید خروجی.')
        return redirect('dashboard:home')


def _generate_csv_export(request, queryset, model_config):
    """تولید خروجی CSV"""
    response = HttpResponse(
        content_type='text/csv; charset=utf-8-sig'
    )
    filename = (
        f"{model_config['filename']}_"
        f"{timezone.now().strftime('%Y%m%d')}.csv"
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{filename}"'
    )

    writer = csv.writer(response)

    # ─── هدرها ───
    writer.writerow(model_config['headers'])

    # ─── داده‌ها ───
    for obj in queryset:
        row = []
        for field in model_config['fields']:
            value = _get_nested_field(obj, field)
            row.append(value)
        writer.writerow(row)

    return response


def _generate_excel_export(request, queryset, model_config):
    """تولید خروجی Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        # ✅ FIX ۳.۸.۱: اطلاع‌رسانی به کاربر
        messages.warning(
            request,
            'کتابخانه Excel (openpyxl) نصب نیست. '
            'خروجی به فرمت CSV تولید می‌شود.'
        )
        return _generate_csv_export(request, queryset, model_config)

    wb = Workbook()
    ws = wb.active
    ws.title = model_config['filename']

    # ─── هدرها با استایل ───
    header_fill = PatternFill(
        start_color='A88B7D',
        end_color='A88B7D',
        fill_type='solid',
    )
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(model_config['headers'], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # ─── داده‌ها ───
    for row_idx, obj in enumerate(queryset, 2):
        for col_idx, field in enumerate(model_config['fields'], 1):
            value = _get_nested_field(obj, field)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # ─── تنظیم عرض ستون‌ها ───
    for col in range(1, len(model_config['headers']) + 1):
        max_length = len(model_config['headers'][col - 1])
        for row in range(2, min(ws.max_row + 1, 100)):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col)].width = adjusted_width

    # ─── ذخیره در HttpResponse ───
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            'application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        ),
    )
    filename = (
        f"{model_config['filename']}_"
        f"{timezone.now().strftime('%Y%m%d')}.xlsx"
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{filename}"'
    )
    return response


def _get_nested_field(obj, field):
    """
    دریافت مقدار فیلد (با پشتیبانی از روابط)
    ✅ FIX ۳.۸.۳: هندل کامل خطاها
    """
    try:
        if '__' in field:
            parts = field.split('__')
            value = obj
            for part in parts:
                if value is None:
                    return ''
                value = getattr(value, part, None)
            return value if value is not None else ''
        result = getattr(obj, field, '')
        return result if result is not None else ''
    except AttributeError:
        logger.warning(
            f"Export field '{field}' not found on "
            f"{obj.__class__.__name__}"
        )
        return ''
    except Exception as e:
        logger.error(
            f"Export field error for '{field}': {e}"
        )
        return ''