"""
سرویس تولید گزارشات (Excel/CSV)
✅ بهینه‌شده: استفاده از iterator() و only()
"""
import io
import csv
from datetime import datetime
from django.core.files.base import ContentFile
from django.utils import timezone
from django.db.models import Sum, Count, Avg
from apps.advanced.models import Report


class ReportService:
    """سرویس تولید گزارشات"""

    @classmethod
    def create_report(cls, user, report_type, format_type, filters):
        """ایجاد درخواست گزارش"""
        report = Report.objects.create(
            user=user,
            report_type=report_type,
            format=format_type,
            filters=filters,
        )
        cls._generate_report(report)
        return report

    @classmethod
    def _generate_report(cls, report):
        """تولید فایل گزارش"""
        try:
            if report.report_type == Report.Type.TRANSACTIONS:
                data = cls._get_transactions_data(report.user, report.filters)
            elif report.report_type == Report.Type.APPOINTMENTS:
                data = cls._get_appointments_data(report.user, report.filters)
            elif report.report_type == Report.Type.REVIEWS:
                data = cls._get_reviews_data(report.user, report.filters)
            else:
                raise ValueError(f'نوع گزارش پشتیبانی نمی‌شود: {report.report_type}')

            if report.format == Report.Format.CSV:
                file_content = cls._generate_csv(data)
                filename = f'report_{report.report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            else:
                file_content = cls._generate_excel(data, report.report_type)
                filename = f'report_{report.report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

            report.file.save(filename, ContentFile(file_content))
            report.records_count = len(data)
            report.file_size = len(file_content)
            report.is_ready = True
            report.completed_at = timezone.now()
            report.expires_at = timezone.now() + timezone.timedelta(days=7)
            report.save()
        except Exception as e:
            report.error_message = str(e)
            report.save()

    @classmethod
    def _get_transactions_data(cls, user, filters):
        """
        ✅ بهینه: استفاده از iterator() و only()
        """
        from apps.payments.models import Transaction

        qs = Transaction.objects.filter(business__owner=user)

        if filters.get('date_from'):
            qs = qs.filter(created_at__date__gte=filters['date_from'])
        if filters.get('date_to'):
            qs = qs.filter(created_at__date__lte=filters['date_to'])
        if filters.get('status'):
            qs = qs.filter(status=filters['status'])
        if filters.get('type'):
            qs = qs.filter(type=filters['type'])

        qs = qs.select_related(
            'user', 'business', 'appointment'
        ).only(
            'tracking_code', 'created_at', 'amount',
            'commission_amount', 'net_amount', 'status', 'type', 'gateway',
            'user__full_name', 'user__phone',
        ).order_by('-created_at')

        # ✅ استفاده از iterator برای کاهش memory
        results = []
        for tx in qs.iterator(chunk_size=1000):
            results.append({
                'کد پیگیری': tx.tracking_code,
                'تاریخ': tx.created_at.strftime('%Y/%m/%d %H:%M'),
                'نام مشتری': tx.user.full_name or tx.user.phone,
                'مبلغ (تومان)': tx.amount,
                'کارمزد (تومان)': tx.commission_amount,
                'مبلغ خالص (تومان)': tx.net_amount,
                'وضعیت': tx.get_status_display(),
                'نوع': tx.get_type_display(),
                'درگاه': tx.gateway or '—',
            })
        return results

    @classmethod
    def _get_appointments_data(cls, user, filters):
        """دریافت داده‌های نوبت‌ها"""
        from apps.bookings.models import Appointment

        qs = Appointment.objects.filter(business__owner=user)

        if filters.get('date_from'):
            qs = qs.filter(date__gte=filters['date_from'])
        if filters.get('date_to'):
            qs = qs.filter(date__lte=filters['date_to'])
        if filters.get('status'):
            qs = qs.filter(status=filters['status'])

        qs = qs.select_related(
            'customer', 'service', 'employee'
        ).only(
            'date', 'time', 'final_price', 'deposit_amount', 'status',
            'customer__full_name', 'customer__phone',
            'service__name',
            'employee__name',
        ).order_by('-date', '-time')

        results = []
        for apt in qs.iterator(chunk_size=1000):
            results.append({
                'تاریخ': apt.date.strftime('%Y/%m/%d'),
                'ساعت': apt.time.strftime('%H:%M'),
                'نام مشتری': apt.customer.full_name or apt.customer.phone,
                'شماره مشتری': apt.customer.phone,
                'خدمت': apt.service.name,
                'کارمند': apt.employee.name if apt.employee else '—',
                'مبلغ کل (تومان)': apt.final_price,
                'بیعانه (تومان)': apt.deposit_amount,
                'وضعیت': apt.get_status_display(),
            })
        return results

    @classmethod
    def _get_reviews_data(cls, user, filters):
        """دریافت داده‌های نظرات"""
        from apps.reviews.models import Review

        qs = Review.objects.filter(business__owner=user)

        if filters.get('date_from'):
            qs = qs.filter(created_at__date__gte=filters['date_from'])
        if filters.get('date_to'):
            qs = qs.filter(created_at__date__lte=filters['date_to'])
        if filters.get('rating'):
            qs = qs.filter(rating=filters['rating'])

        qs = qs.select_related(
            'customer', 'service', 'business'
        ).only(
            'created_at', 'rating', 'comment', 'is_approved',
            'customer__full_name', 'customer__phone',
            'service__name',
        ).order_by('-created_at')

        results = []
        for review in qs.iterator(chunk_size=1000):
            results.append({
                'تاریخ': review.created_at.strftime('%Y/%m/%d'),
                'نام مشتری': review.customer.full_name or review.customer.phone,
                'خدمت': review.service.name if review.service else '—',
                'امتیاز': review.rating,
                'نظر': review.comment,
                'وضعیت': 'تایید شده' if review.is_approved else 'در انتظار',
            })
        return results

    @classmethod
    def _generate_csv(cls, data):
        """تولید CSV"""
        if not data:
            return b''
        output = io.StringIO()
        output.write('\ufeff')
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        return output.getvalue().encode('utf-8-sig')

    @classmethod
    def _generate_excel(cls, data, report_type):
        """تولید Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            return cls._generate_csv(data)

        if not data:
            return b''

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Report'

        header_font = Font(name='Tahoma', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='A88B7D', end_color='A88B7D', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row, record in enumerate(data, 2):
            for col, key in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=record[key])
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in range(1, len(headers) + 1):
            max_length = max(
                len(str(ws.cell(row=row, column=col).value or ''))
                for row in range(1, len(data) + 2)
            )
            ws.column_dimensions[chr(64 + col)].width = min(max_length + 2, 30)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()